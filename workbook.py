# todo create a sheet specific interface
def write(meeting_end_time, time_use, classlist_file):
    import time
    from datetime import datetime
    import win32com.client
    import sql
    # loading workbook
    from openpyxl import load_workbook, styles
    from openpyxl.utils.cell import column_index_from_string, get_column_letter
    # classlist_file = "D:\Attendace Code\DATA\CLASSLIST.xlsx"
    workbook = load_workbook(filename=classlist_file)
    sheet = workbook.active

    # def arranging_columnwidthafter_insert(t):
    #     for i in range(t + 1, empty_column(t + 1)):
    #         column = get_column_letter(i)
    #         sheet.column_dimensions[f'{column}'].width = 22

    def empty_column(i=1):

        while True:
            if sheet.cell(row=1, column=i).value == None:
                break
            i += 1
        return i

    def write_column(meeting_end_time, time_use):
        # we basically try to use number 1 as first occurrence not 0 because of column number
        if empty_column() == 3:
            return get_column_letter(empty_column())
        k = [sheet.cell(row=1, column=i).value for i in range(3, empty_column())]
        code = '%m/%d/%Y, %I:%M:%S %p'
        list_of_times = []
        for x in k:
            list_of_times.append(datetime.strptime(x, code))
        list_of_times.append(datetime.strptime(meeting_end_time, code))
        list_of_times.sort()  # because meeting ending times may not be in order
        t = list_of_times.index(datetime.strptime(meeting_end_time,
                                                  code)) + 3  # because if at first position we want to be it 3 for 3rd column
        # checking if same time entry and if column has previous and updated values
        # till now new column hasn't been written
        # t tells the position of meeting_end_time in list, nothing is written yet in excel
        # first condition checks if not last column
        # after inserting t is the idx of inserted column
        if t - 3 != len(list_of_times) - 1 and list_of_times[t - 3] == list_of_times[t - 2]:
            print('\nNNNN')
            i = 2
            list_o_s = dict(time_use)
            while sheet.cell(row=i, column=t).value is not None:
                if sheet.cell(row=i, column=2).value in list_o_s and \
                        sheet.cell(row=i, column=t).value < list_o_s[sheet.cell(row=i, column=2).value]:
                    sheet.delete_cols(idx=t)  # here the column is deleted and inserted , t is inserted column
                    sheet.insert_cols(idx=t)
                    print('\nYA')
                    return get_column_letter(t)
                i += 1
            return None
        if sheet.cell(column=t, row=1).value is not None:  # this is for new date
            sheet.insert_cols(idx=t)
            # arranging_columnwidthafter_insert(t)
            print('\nNap')
        return get_column_letter(t)  # this is returning if none we need not to insert

    classlist = [i[0] for i in sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True) if i[0] is not None]
    # or   k = [i.value for i in sheet['b']]
    print("I am here 2")
    # todo checking if of same class
    ctr = 0
    for i in range(len(time_use)):
        if time_use[i][0] in classlist:
            ctr += 1
            print(ctr)
        if ctr == 7:
            print('reache')
            break
        elif i == len(time_use) - 1:
            print('not in ')

    # writing to excel
    column = write_column(meeting_end_time, time_use)
    print('I am here column letter got and insertion made:---', column)
    if ctr == 7 and column != None:
        print('\n\n', column, '--COLUMN', '\n\n')
        print('I am here 3')
        sheet.column_dimensions[f'{column}'].width = 22
        sheet[f"{column}" + '1'] = meeting_end_time
        big_red_text = styles.Font(color=styles.colors.RED, size=14)
        center_aligned_text = styles.Alignment(horizontal="center")
        for i in time_use:
            if i[0] in classlist:
                position = classlist.index(i[0]) + 2
                sheet[f"{column}" + f"{position}"] = i[1]
                sheet[f"{column}" + f"{position}"].alignment = center_aligned_text
            # else:
            #     print(i, 'not in classlist')
            #
            #     position = len(classlist) + 2
            #     sheet['A' + f'{position}'] = len(classlist) + 1
            #     sheet['B' + f'{position}'] = i[0]
            #     sheet[f"{column}" + f"{position}"] = i[1]
            #     sheet[f"{column}" + f"{position}"].alignment = center_aligned_text
            #     classlist = [p[0] for p in sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True) \
            #                  if p[0] is not None]
            #
            #     for j in range(3, column_index_from_string(column)):
            #         print(column_index_from_string(column))
            #         sheet.cell(row=position, column=j).value = 'Absent'
            #         print(position, j, '----+++')
            #         sheet.cell(row=position, column=j).font = big_red_text
            #         sheet.cell(row=position, column=j).alignment = center_aligned_text
            #     r = column_index_from_string(column) + 1
            #     while sheet.cell(row=position - 1, column=r).value is not None \
            #             and sheet.cell(row=position, column=r).value is None:
            #         print(position, r, '----+++')
            #         sheet.cell(row=position, column=r).value = 'Absent'
            #         sheet.cell(row=position, column=r).font = big_red_text
            #         sheet.cell(row=position, column=r).alignment = center_aligned_text
            #         r += 1

        l = [f.value for f in sheet[f'{column}']]
        rt = [f.value for f in sheet['B']]
        for m in range(len(l)):
            if l[m] is None and rt[m] is not None:
                sheet[f"{column}" + f"{m + 1}"] = 'Absent'
                sheet[f"{column}" + f"{m + 1}"].font = big_red_text
                sheet[f"{column}" + f"{m + 1}"].alignment = center_aligned_text
    classlist.sort()
    # required when student not in classlist
    # for i in range(2, len(classlist) + 2):
    #     temp = classlist.index(sheet['B' + f'{i}'].value)
    #     sheet['A' + f'{i}'] = temp + 1
    print('I am here 4')
    time.sleep(0.5)
    if ctr == 7:
        # sql.table_creation(meeting_end_time, time_use, sql.name_of_database(classlist_file))
        workbook.save(filename=classlist_file)
    workbook.close()
    print('I am here 5')
    if ctr == 7:
        try:
            global excel
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            global wb

            wb = excel.Workbooks.Open(classlist_file)
            ws = wb.Worksheets(wb.Sheets(1).Name)

            ws.Range(f'A2:{get_column_letter(empty_column() - 1)}{len(classlist) + 1}').Sort(Key1=ws.Range('B1'),
                                                                                             Order1=1,
                                                                                             Orientation=1)
            ws.Activate()
            ws.Range("C2").Select()  # add the () at the end here
            excel.ActiveWindow.FreezePanes = True
            ws.Range(f"{get_column_letter(empty_column() - 1)}2").Select()  # so that i don't have to scroll
            ws.Columns.AutoFit()
        except Exception as e:
            print(e)
        finally:
            wb.Save()
            excel.Application.Quit()


def making_classlist(time_use, classlist_file):
    # loading workbook
    import win32com.client
    from openpyxl import Workbook
    workbook = Workbook()
    sheet1 = workbook.active
    # sheet1.column_dimensions['B'].width = max([len(i) for i in time_use]) + 4
    # sheet1.column_dimensions['A'].width = len(' S.No ')
    sheet1['A1'] = 'S.No'
    sheet1['B1'] = 'Name'
    print('i am working on classlist')
    # deleting names
    i = 0
    for i in range(len(time_use)):
        sheet1['A' + f'{i + 2}'] = i + 1
        sheet1['B' + f'{i + 2}'] = time_use[i]
    workbook.save(filename=classlist_file)
    try:
        global excel
        excel = win32com.client.Dispatch('Excel.Application')
        excel.Visible = False
        global wb
        wb = excel.Workbooks.Open(classlist_file)
        ws = wb.Worksheets(wb.Sheets(1).Name)
        ws.Columns.AutoFit()
        ws.Activate()
        ws.Range("C2").Select()  # add the () at the end here
        excel.ActiveWindow.FreezePanes = True
    except Exception as e:
        print(e)
    finally:
        wb.Save()
        excel.Application.Quit()

    print('classlist made')
